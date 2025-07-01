import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl import Workbook
import requests
from datetime import datetime

# 从Secrets获取Airtable配置（需用户在Streamlit社区云或本地secrets.toml中配置）
AIRTABLE_API_KEY = st.secrets.get("airtable_api_key", "")
AIRTABLE_BASE_ID = st.secrets.get("airtable_base_id", "")
AIRTABLE_TABLE_NAME = "tblWyetzvY1weZjDv"


def get_airtable_data():
    if not AIRTABLE_API_KEY or not AIRTABLE_BASE_ID:
        st.error("请配置Airtable API密钥和基ID（通过Streamlit Secrets或环境变量）")
        return pd.DataFrame()
    
    url = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{AIRTABLE_TABLE_NAME}"
    headers = {"Authorization": f"Bearer {AIRTABLE_API_KEY}"}
    
    all_records = []
    offset = None
    
    # Airtable API分页获取所有数据
    while True:
        params = {"offset": offset} if offset else {}
        try:
            response = requests.get(url, headers=headers, params=params)
            response.raise_for_status()
            data = response.json()
            all_records.extend(data.get("records", []))
            offset = data.get("offset")
            if not offset:
                break
        except Exception as e:
            st.error(f"获取Airtable数据失败: {str(e)}")
            return pd.DataFrame()
    
    records = []
    for record in all_records:
        fields = record.get("fields", {})
        date_str = fields.get("日期")  # 假设Airtable中有"日期"字段（格式YYYY-MM-DD）
        if date_str:
            try:
                month = datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y-%m")
            except ValueError:
                month = "日期格式错误"
        else:
            month = "未记录日期"
        
        records.append({
            "月份": month,
            "患者姓名": fields.get("患者姓名", ""),  # 使用Airtable实际字段名"name"获取患者姓名
            "科室": fields.get("科室", ""),          # 添加科室字段
            "病例类型": fields.get("业务类型", "未知类型"),
            "分值": fields.get("分值", 0),
            "备注": fields.get("备注", "")           # 添加备注字段
        })
    
    return pd.DataFrame(records)


# 主应用
st.title("📊 病例制作情况与分值分析看板")

# 获取并缓存数据（每小时刷新）
@st.cache_data(ttl=3600)
def cached_airtable_data():
    return get_airtable_data()
df = cached_airtable_data()

if not df.empty:
    # 数据预处理
    monthly_group = df.groupby(["月份", "病例类型"]).agg({
        "分值": ["sum", "count"]  # 计算每月每种类型的总分和病例数
    }).reset_index()
    monthly_group.columns = ["月份", "病例类型", "总分值", "病例数"]
    
    monthly_total = df.groupby("月份").agg({
        "分值": ["sum", "count"]
    }).reset_index()
    monthly_total.columns = ["月份", "月度总分", "病例总数"]
    
    # 侧边栏筛选
    st.sidebar.header("筛选条件")
    selected_months = st.sidebar.multiselect("选择月份", df["月份"].unique(), default=df["月份"].unique())
    filtered_group = monthly_group[monthly_group["月份"].isin(selected_months)]
    filtered_total = monthly_total[monthly_total["月份"].isin(selected_months)]

    # 主内容展示
    st.subheader("📅 月度病例类型分布（堆叠柱状图）")
    st.bar_chart(filtered_group, x="月份", y="总分值", color="病例类型", use_container_width=True)
    
    st.subheader("📈 月度总分趋势")
    st.line_chart(filtered_total, x="月份", y=["月度总分", "病例总数"], use_container_width=True)
    
    st.subheader("📋 详细数据表格")
    st.dataframe(filtered_group, use_container_width=True)
    # 添加Excel导出功能
    st.subheader("💾 导出月度数据")
    export_month = st.selectbox("选择导出月份", sorted(df["月份"].unique()))
    export_data = df[df["月份"] == export_month].copy()

    # 调整列顺序和名称以匹配Excel格式
    if not export_data.empty:
        # 保留所有列并仅重命名需要的字段
        # 移除分值重命名，避免列名冲突
        export_data = export_data.rename(columns={
            "病例类型": "业务类型"
        })

        # 生成Excel文件
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "病例统计"
        
        # 定义必要的列名列表
        required_columns = ["患者姓名", "科室", "病例类型", "分值", "备注"]
        
        column_mapping = {
            "患者姓名": "患者姓名",  # Airtable字段"name"映射为Excel列"患者姓名"
            "科室": "科室",        # 科室字段映射
            "业务类型": "病例类型",  # Airtable字段"业务类型"映射为Excel列"病例类型"
            "分值": "分值",
            "备注": "备注"         # 备注字段ID
        }
        export_data = export_data.rename(columns=column_mapping)
        
        # 确保所有必要列存在
        for col in required_columns:
            if col not in export_data.columns:
                export_data[col] = ""
        
        # 按要求顺序排列列
        export_data = export_data[required_columns]
        
        # 添加序号列作为第一列
        export_data.insert(0, "序号", range(1, len(export_data) + 1))
        
        # 写入表头
        for col_num, column_title in enumerate(export_data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        
        # 写入数据行
        for row_num, row_data in enumerate(export_data.itertuples(index=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # 添加总分行
        total_row = len(export_data) + 2  # 最终确认：数据从第2行开始，总计行在数据行后显示（无空白行）
        ws.cell(row=total_row, column=4, value="总计")  # 将总计标签移至业务类型列
        ws.cell(row=total_row, column=5, value=export_data["分值"].sum())  # 修正总分列位置
        
        wb.save(output)
        output.seek(0)

        # 下载按钮
        # 格式化月份为不带前导零的格式
        year, month = export_month.split('-')
        formatted_month = f"{int(month)}月份"
        file_name = f"任彬彬{year}年{formatted_month}病例统计.xlsx"
        st.download_button(
            label=f"下载{year}年{formatted_month}数据",
            data=output,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
 
else:
    st.info("请检查Airtable配置或确保表中存在数据")
